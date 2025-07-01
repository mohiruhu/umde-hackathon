"""
rule_chunk_parser.py

Unified parser module to extract rule text chunks from both PDF and XLSX CMS source documents.
This version includes:
- Dynamic header matching for CEM 837P/I/DME Excel files.
- Buffered parsing from CMS PDF documents (e.g. PCUG, EDPS).
"""

from typing import List, Dict, Any
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import pdfplumber
import re
import logging

logger = logging.getLogger("cms_parser")


def parse_xlsx_rules(xlsx_path: str) -> List[Dict[str, Any]]:
    results: List[Dict[str, Any]] = []
    skipped_rows = 0

    try:
        logger.info(f"Opening XLSX file: {xlsx_path}")
        workbook = load_workbook(xlsx_path, data_only=True)
        if not workbook.sheetnames:
            logger.warning(f"No sheets found in XLSX file: {xlsx_path}")
            return []

        sheet = workbook.active
        if not sheet:
            logger.error(f"No active sheet found in workbook: {xlsx_path}")
            raise ValueError("No active sheet found in workbook.")

        expected_headers: Dict[str, List[str]] = {
            "rule_id": ["edit reference", "reference id", "837p edit reference", "837i edit reference", "cedi edit reference"],
            "description": ["description"]
        }

        header_map: Dict[str, int] = {}
        header_row_index = -1

        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            cells = [str(cell).strip().lower() if cell else "" for cell in row]
            for idx, value in enumerate(cells):
                for key, patterns in expected_headers.items():
                    if any(p in value for p in patterns):
                        if key not in header_map:
                            header_map[key] = idx
            if "rule_id" in header_map and "description" in header_map:
                header_row_index = i
                logger.info(f"Detected header row at line {header_row_index}")
                break
        else:
            logger.error("Required headers not found in XLSX file.")
            raise ValueError("Required headers not found in XLSX file.")

        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if i <= header_row_index:
                continue

            cells = [str(cell).strip() if cell else "" for cell in row]
            try:
                rule_id = cells[header_map["rule_id"]]
                description = cells[header_map["description"]]
            except IndexError:
                skipped_rows += 1
                logger.warning(f"Skipping row {i}: index error in data structure.")
                continue

            if not rule_id or not description:
                skipped_rows += 1
                logger.warning(f"Skipping row {i} with rule_id={rule_id}: missing rule_id or description.")
                continue

            results.append({
                "rule_id": rule_id,
                "raw_row": description,
                "source_type": "xlsx",
                "doc": Path(xlsx_path).stem,
                "meta": {
                    "row": i,
                    "sheet": sheet.title
                }
            })

        logger.info(f"Parsed {len(results)} rules from {sheet.title}; Skipped {skipped_rows} rows.")
        return results

    except FileNotFoundError:
        logger.error(f"File not found: {xlsx_path}")
        raise

    except InvalidFileException:
        logger.error(f"Invalid Excel file format: {xlsx_path}")
        raise

    except Exception as e:
        logger.error(f"Unexpected error while parsing XLSX: {e}")
        raise


def parse_pdf_rules(pdf_path: str, start_page: int, end_page: int) -> List[Dict[str, Any]]:
    rule_start_pattern = re.compile(r"^(?P<code>\d{3})[-:]?\s+(?P<type>[RAI])\s+(?P<title>.+)", re.IGNORECASE)
    extracted: List[Dict[str, Any]] = []
    seen_trcs: set[str] = set()
    buffer: List[str] = []
    current_trc = ""
    current_title = ""
    current_type = ""
    page_range = range(start_page - 1, end_page)

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for i in page_range:
                if i >= len(pdf.pages):
                    logger.warning(f"Skipping page {i + 1}: beyond document length.")
                    continue

                page = pdf.pages[i]
                text = page.extract_text() or ""
                lines = text.splitlines()
                for line in lines:
                    match = rule_start_pattern.match(line.strip())
                    if match:
                        if buffer and current_trc:
                            extracted.append({
                                "rule_id": f"TRC{current_trc}",
                                "short_definition": current_title,
                                "definition": " ".join(buffer).strip(),
                                "source_type": "pdf",
                                "doc": Path(pdf_path).stem,
                                "meta": {
                                    "page": i + 1,
                                    "type": current_type
                                }
                            })
                            buffer = []
                        current_trc = match.group("code")
                        if current_trc in seen_trcs:
                            continue
                        seen_trcs.add(current_trc)
                        current_type = match.group("type")
                        current_title = match.group("title")
                    elif current_trc:
                        buffer.append(line.strip())

            if buffer and current_trc:
                extracted.append({
                    "rule_id": f"TRC{current_trc}",
                    "short_definition": current_title,
                    "definition": " ".join(buffer).strip(),
                    "source_type": "pdf",
                    "doc": Path(pdf_path).stem,
                    "meta": {
                        "page": end_page,
                        "type": current_type
                    }
                })

        logger.info(f"Parsed {len(extracted)} TRC rules from PDF: {pdf_path}")
        return extracted

    except FileNotFoundError:
        logger.error(f"PDF file not found: {pdf_path}")
        raise

    except Exception as e:
        logger.error(f"Unexpected error parsing PDF: {e}")
        raise